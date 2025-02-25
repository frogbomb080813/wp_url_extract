from flask import Flask, request, render_template, redirect, url_for, send_file
import requests
from bs4 import BeautifulSoup
import openpyxl
import os

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get the number of target URLs from the form
        num_targets = int(request.form['num_targets'])
        
        # Get the target URLs from the form
        target_urls = [request.form[f'target_url_{i+1}'] for i in range(num_targets)]
        
        # Initialize the Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Extracted Data"
        sheet["A1"] = "Target URL"
        sheet["B1"] = "Title"
        sheet["C1"] = "Video URL"
        sheet["D1"] = "Thumbnail URL"

        # Loop through each target URL and extract the required data
        for row, target_url in enumerate(target_urls, start=2):
            # Send an HTTP GET request to the target URL
            response = requests.get(target_url)
            
            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract the title of the page
            title_tag = soup.find('title')
            title = title_tag.text if title_tag else None

            # Extract the video URL using the meta tag with itemprop="contentURL"
            meta_tag = soup.find('meta', {'itemprop': 'contentURL'})
            video_url = meta_tag['content'] if meta_tag else None
            
            # Extract the thumbnail URL using the meta tag with itemprop="thumbnailUrl"
            thumbnail_meta_tag = soup.find('meta', {'itemprop': 'thumbnailUrl'})
            thumbnail_url = thumbnail_meta_tag['content'] if thumbnail_meta_tag else None
            
            # Save the extracted data to the Excel sheet
            sheet[f"A{row}"] = target_url
            sheet[f"B{row}"] = title
            sheet[f"C{row}"] = video_url
            sheet[f"D{row}"] = thumbnail_url
        
        # Save the Excel file
        workbook.save("extracted_data.xlsx")
        
        # Redirect to the results page
        return redirect(url_for('results'))
    
    return render_template('index.html')

@app.route('/results')
def results():
    # Load the Excel file and extract the data
    workbook = openpyxl.load_workbook("extracted_data.xlsx")
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    
    return render_template('/results.html', data=data)

@app.route('/download')
def download():
    path = "extracted_data.xlsx"
    return send_file(path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
